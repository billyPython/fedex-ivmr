import argparse
import codecs
import csv
from math import ceil

import openpyxl
from openpyxl.drawing import image
from xlsxwriter import Workbook

fieldnames = [
    'Date',
    # 'Date End of Day',
    'State Code',
    'Beginning Odometer',
    'Ending Odometer',
    # 'LCVs',
    # 'Origin/Destination/City\n'
    # 'Spot Number/City',
    # 'Highway or Route\nTraveled'
]

WRITE_FILE = 'INDIVIDUAL VEHICLE MILEAGE RECORD (IVMR).csv'
SAVE_TO_FILE = 'ivmr_report.xls'
TEMPLATE_FILE = 'IVMR.xlsx'
TEMPLATE_FILE_IMAGE = 'fedex.png'

ivmr_field_mapping = {
    'Vehicle ID': 'D8',
    'Date': 'C',
    'State Code': 'F',
    'Beginning Odometer': 'I',
    'Ending Odometer': 'L'
}

#TODO: Run this as class!! Restructure everything !

def cleanup(data):
    for d in data:
        for k, v in d.items():
            if k not in fieldnames:
                d.pop(k)

def load_data(path=None, file=None):
    """
    :param path: Path to the CSV file we are gonna read and use
    :param file: Already red CVS file with csv.DictReader
    :return: Data in list of dictionaries
    """
    data = []

    if path:
        csv_read = codecs.open(path, 'rU', 'utf-16')
        csv_records = csv.DictReader(csv_read)
    elif file:
        csv_records = file
    else:
        raise Exception("There is no input!")

    for records in csv_records:
        k, v = records.items()[0]
        load = zip(k.split('\t'), v.split('\t'))
        data.append(dict(load))

    return data

def make_xls():
    workbook = Workbook(WRITE_FILE[:-4] + '.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.set_column('A:B', 10)
    worksheet.set_column('B:E', 18)

    with codecs.open(WRITE_FILE, 'r', 'utf-8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)
        workbook.close()

        print('Your file is ---{}---!'.format(WRITE_FILE))
        print("Done!")

def make_csv_ivmr_report(data):
    """
    :param data: List of dictionaries
    This function parses `data` and saves it into new CSV file.
    """
    csv_write = open(WRITE_FILE, 'wb')

    # write vehicle code on start
    vehicle_id = data[0]['Vehicle ID']
    writer = csv.writer(csv_write, delimiter=':')

    writer.writerow(['Home Station#', '9072'])
    writer.writerow([])

    writer.writerow(['Vehicle ID', vehicle_id])
    writer.writerow([])

    tmp = int(ceil(float(data[0]['Beginning Odometer'])))
    for d in data:
        d['Date'] = d['Date'].split(' ')[0]
        d['Beginning Odometer'] = tmp
        d['Ending Odometer'] = tmp + int(ceil(float(d['Total State Miles'])))
        tmp = d['Ending Odometer']

    cleanup(data)

    writer = csv.DictWriter(csv_write, fieldnames=fieldnames)
    writer.writeheader()
    for d in data:
        writer.writerow(d)

def make_xls_ivmr_report(data=None):
    """
    :param data: List of dictionaries
    This function parses `data` and saves it into `TEMPLATE_FILE` .xls file
    """
    wb = openpyxl.load_workbook(TEMPLATE_FILE, keep_vba=True)
    # wb.template = True
    ws = wb.active
    ws[ivmr_field_mapping['Vehicle ID']].value = data[0]['Vehicle ID']

    tmp = int(ceil(float(data[0]['Beginning Odometer'])))
    for i, d in enumerate(data, 22):

        d['Beginning Odometer'] = tmp
        d['Ending Odometer'] = tmp + int(ceil(float(d['Total State Miles'])))

        #Writing data to IVMR.xls cells
        ws['C{}'.format(i)] = d['Date'].split(' ')[0]
        ws['F{}'.format(i)] = d['State Code']
        ws['I{}'.format(i)] = tmp
        ws['L{}'.format(i)] = d['Ending Odometer']

        tmp = d['Ending Odometer']

    img = image.Image(TEMPLATE_FILE_IMAGE)
    img.height = 96
    img.width = 230
    ws.add_image(img, anchor='A1')

    wb.save(SAVE_TO_FILE)

    print ('Done!')
    print ('Your file is ---{}---!'.format(SAVE_TO_FILE))

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--p",  nargs='+', help="Path to the file.")
    parser.add_argument("--fn", nargs='+', help="Set filename.")
    args = parser.parse_args()
    if args.fn:
        WRITE_FILE = args.fn[0]
    if len(args.p) > 1:
        raise Exception('Put your file path under quotation marks --> "file path"')

    data = load_data(path=args.p[0]) # Load data from CSV to dictionary
    make_csv_ivmr_report(data)
    make_xls()
    # make_xls_ivmr_report(data=data)